from src.Models.Users import User
from src.Models.Participants import Participant
from dotenv import load_dotenv
from os import environ
from openpyxl import Workbook


load_dotenv()


class GeneratorReport:
    def __init__(self, event, attendance, users_not_attend):
        self.event = event
        self.attendance = attendance
        self.users_not_attend = users_not_attend

    def generate_report(self):
        # Crear un nuevo libro de trabajo de Excel
        wb = Workbook()

        # Crear hojas de trabajo para los asistentes y los no asistentes
        ws_attend = wb.create_sheet("Asistentes")
        ws_not_attend = wb.create_sheet("No Asistentes")

        # Agregar encabezados a las hojas de trabajo
        headers = ["user_id", "name", "email", "identification document", "phone", "join_code", "registration_date", "access_code"]
        ws_attend.append(headers)
        ws_not_attend.append(headers)

        # Agregar datos de asistentes a la hoja de trabajo correspondiente
        for participant in self.attendance:
            user = User.query.get(participant.user_id)
            if user:
                participant_data = Participant.query.filter_by(user_id=user.user_id, event_id=self.event.event_id).first()
                if participant_data:
                    ws_attend.append([user.user_id, user.name, user.email, user.identification_document, user.phone, participant_data.join_code, participant.registration_date, self.event.access_code])

        # Agregar datos de no asistentes a la hoja de trabajo correspondiente
        for user_id in self.users_not_attend:
            user = User.query.get(user_id)
            if user:
                participant_data = Participant.query.filter_by(user_id=user.user_id, event_id=self.event.event_id).first()
                if participant_data:
                    ws_not_attend.append([user.user_id, user.name, user.email, user.identification_document, user.phone, None, None, self.event.access_code])

        # Definir la ruta donde guardar el archivo de reporte
        report_path = f"{environ.get('PATH_REPORTS_ASSISTANCE')}\{self.event.name}_{self.event.event_id}.xlsx"
        
        # Guardar el libro de trabajo en la ruta especificada
        wb.save(report_path)

        return report_path
